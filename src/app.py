"""
機材レンタル管理 - Flaskアプリケーション

商品コードを軸とした機材レンタル管理のデスクトップアプリケーション。
PyWebViewの専用ウィンドウで表示し、ウィンドウを閉じるとプロセスも終了する。
"""

import os
import sys
import threading

from dotenv import load_dotenv

# .exe 化した場合は実行ファイルと同じディレクトリの .env を読む
_dotenv_path = os.path.join(
    os.path.dirname(sys.executable) if getattr(sys, 'frozen', False)
    else os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    '.env'
)
load_dotenv(_dotenv_path)

# pythonw.exe 等で標準出力が無い環境でのクラッシュ（sys.stdout is None）を防ぐ
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w')
if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')

from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from database import initialize_database, get_all_equipment, sync_equipment_from_list, \
    search_by_product_code, register_rental, process_return
from azure_ocr import analyze_image


# ---------- パス解決 ----------
# PyInstaller で .exe 化した場合、リソースは sys._MEIPASS 配下に展開される。
# data/ は永続データのため、.exe と同じディレクトリに配置する。
def _get_resource_dir() -> str:
    """templates/static 等のリソースディレクトリを返す。"""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _get_app_dir() -> str:
    """data/ 等の永続データを配置するディレクトリを返す。"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


RESOURCE_DIR = _get_resource_dir()
APP_DIR = _get_app_dir()

app = Flask(
    __name__,
    template_folder=os.path.join(RESOURCE_DIR, 'templates'),
    static_folder=os.path.join(RESOURCE_DIR, 'static')
)

# Excelアップロードの一時保存先
UPLOAD_DIR = os.path.join(APP_DIR, "data", "uploads")
ALLOWED_EXTENSIONS = {".xlsx"}


def is_allowed_file(filename: str) -> bool:
    """許可された拡張子かチェックする。"""
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_EXTENSIONS


# ---------- 画面表示 ----------

@app.route("/")
def index():
    """メイン画面（検索・レンタル・返却を1画面で操作）"""
    return render_template("index.html")


@app.route("/equipment")
def equipment_page():
    """機材マスタ管理画面"""
    return render_template("equipment.html")


# ---------- API ----------

@app.route("/api/search")
def api_search():
    """商品コードで検索。機材情報・レンタル状態・履歴を返す。"""
    product_code = request.args.get("product_code", "").strip()
    if not product_code:
        return jsonify({"error": "商品コードを入力してください"}), 400

    result = search_by_product_code(product_code)
    if result is None:
        return jsonify({"error": "該当する商品コードが見つかりません"}), 404

    return jsonify(result)


@app.route("/api/rental", methods=["POST"])
def api_rental():
    """レンタル登録。"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "リクエストデータが不正です"}), 400

    required_fields = ["product_code", "borrower_name", "rental_start", "rental_end"]
    for field in required_fields:
        if not data.get(field, "").strip():
            return jsonify({"error": f"{field} は必須です"}), 400

    result = register_rental(
        product_code=data["product_code"].strip(),
        borrower_name=data["borrower_name"].strip(),
        rental_start=data["rental_start"].strip(),
        rental_end=data["rental_end"].strip()
    )

    if result["success"]:
        return jsonify(result)
    else:
        return jsonify(result), 409


@app.route("/api/return", methods=["POST"])
def api_return():
    """返却処理。"""
    data = request.get_json()
    if not data or not data.get("product_code", "").strip():
        return jsonify({"error": "商品コードが必要です"}), 400

    result = process_return(data["product_code"].strip())

    if result["success"]:
        return jsonify(result)
    else:
        return jsonify(result), 409


@app.route("/api/analyze-image", methods=["POST"])
def api_analyze_image():
    """画像からレンタル情報を解析する。機材がレンタル中の場合は現在の状況も返す。"""
    ALLOWED_CONTENT_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/bmp", "image/tiff"}

    if "image" not in request.files:
        return jsonify({"error": "画像ファイルが必要です"}), 400

    image_file = request.files["image"]
    content_type = image_file.content_type or "image/png"
    if content_type not in ALLOWED_CONTENT_TYPES:
        return jsonify({"error": "対応形式: PNG, JPEG, BMP, TIFF"}), 400

    image_bytes = image_file.read()

    try:
        extracted = analyze_image(image_bytes, content_type)
    except ValueError as e:
        return jsonify({"error": str(e)}), 500
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 502
    except Exception as e:
        return jsonify({"error": f"画像解析に失敗しました: {str(e)}"}), 500

    # 機材IDが取れた場合、現在のレンタル状況も確認する
    rental_status = None
    if extracted.get("product_code"):
        rental_status = search_by_product_code(extracted["product_code"])

    return jsonify({
        "extracted": extracted,
        "rental_status": rental_status
    })


@app.route("/api/equipment")
def api_equipment_list():
    """機材マスタ一覧を返す。"""
    return jsonify(get_all_equipment())


@app.route("/api/equipment/import", methods=["POST"])
def api_equipment_import():
    """Excelファイルから機材マスタを同期インポートする。"""
    if "file" not in request.files:
        return jsonify({"error": "ファイルが選択されていません"}), 400

    file = request.files["file"]
    if not file.filename or not is_allowed_file(file.filename):
        return jsonify({"error": ".xlsx ファイルのみアップロード可能です"}), 400

    # 一時ファイルとして保存
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, "import.xlsx")
    file.save(filepath)

    try:
        workbook = load_workbook(filepath, read_only=True)
        sheet = workbook.active

        items = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 列数が不足する行はスキップ
            if len(row) < 2:
                continue
            # 1列目: 商品コード, 2列目: 商品名
            if row[0] is None or row[1] is None:
                continue
            product_code = str(row[0]).strip()
            equipment_name = str(row[1]).strip()
            if product_code and equipment_name:
                items.append({"product_code": product_code, "equipment_name": equipment_name})

        workbook.close()

        if not items:
            return jsonify({"error": "有効なデータが見つかりません。1列目に商品コード、2列目に商品名を入力してください"}), 400

        result = sync_equipment_from_list(items)
        return jsonify(result)

    except Exception as e:
        return jsonify({"error": f"ファイルの読み込みに失敗しました: {str(e)}"}), 500
    finally:
        # 一時ファイル削除
        if os.path.exists(filepath):
            os.remove(filepath)


# ---------- 起動 ----------

def _start_flask_server() -> None:
    """Flask サーバーをバックグラウンドスレッドで起動する。"""
    app.run(debug=False, port=5000, use_reloader=False)


if __name__ == "__main__":
    initialize_database()

    # Flask をバックグラウンドスレッドで起動（PyWebView がメインスレッドを占有するため）
    server_thread = threading.Thread(target=_start_flask_server, daemon=True)
    server_thread.start()

    # PyWebView で専用ウィンドウを表示
    # ウィンドウを閉じるとプロセスが自然に終了する
    import webview
    webview.create_window(
        title="機材レンタル管理",
        url="http://127.0.0.1:5000",
        width=900,
        height=700,
        min_size=(600, 400),
        maximized=True,
        text_select=True
    )
    webview.start()
